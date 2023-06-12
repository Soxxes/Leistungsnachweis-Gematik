import os
import sys
import logging

import openpyxl
import pandas as pd


column_names = [
        "Entry Date", "First Name",	"Last Name", "Email", "Client Name", "Client Code",
        "Project Name",	"Project Code",	"Task Name", "Task Code", "Hours", "Comments",
        "Supervisor Name (Current)", "Time Off Type", "Work Type (AUT)", "Work Type (DEU)",
        "Work Type (CHE)", "Work Location", "Time Entry Approval Status", "Timesheet Approval Status"
    ]


def get_input(cwd):
    try:
        dirs = os.listdir(cwd)
        # check if there are mutiple Replicon Exports in the current directory
        count = 0
        for fname in dirs:
            if "Timesheet Hours" in fname:
                count += 1
                _fname = fname
        # if not, choose the one
        if count == 1:
            file = _fname
        # if multiple, let user choose
        elif count > 1:
            replicon_files = [dirname for dirname in dirs if "Timesheet Hours" in dirname]
            file_map = {i: dirname for i, dirname in enumerate(replicon_files, 1)}
            print("Found the following Replicon Exports:")
            for i, dirname in file_map.items():
                print(f"{i} - {dirname}")

            for _ in range(3):
                choice = input("\nSelect your file: ")
                logging.info(f"User's choice in 'get_input' function: {choice}")
                try:
                    file = file_map[int(choice)]
                except ValueError:
                    print("[WARNING] Please provide a number.\n")
                    continue
                except KeyError:
                    print("[WARNING] Please choose one of the numbers displayed. \n")
                    continue
                else:
                    break
            else:
                logging.warning("User entered three invald inputs. Script terminates.")
                input("Received three invalid inputs. Script termintates. Please press any key ...")
                sys.exit()

        # if none, exit
        else:
            logging.warning("No Replicon Export provided. Script terminates.")
            input("No Replicon Export found. Press any key to exit ...")
            sys.exit()

        logging.info("Successfully selected input file.")
        return file

    except Exception as e:
        logging.info("Error in 'get_input' function. Terminated with error:")
        logging.error(f"{e}")
        sys.exit()

def handle_failed_input():
    print("Ops ... Seems like your Replicon Export has column names in a different language than English.")
    print("You can fix that by exporting your timesheet again or rename the column names.\n")
    column_names_wb = openpyxl.Workbook()
    column_names_ws = column_names_wb.active
    for col, name in enumerate(column_names, 1):
        column_names_ws.cell(row=1, column=col).value = name
    column_names_wb.save("column_names.xlsx")
    print("There should be a file named 'column_names.xlsx' with the correct column names.")
    input("Please change your column names (you can copy and paste them) and re-run the program.\n"\
          + "Press any key to exit ...")
    logging.warning("Script terminates. But correct column names were given to the user.")
    sys.exit()

def select_client(config: dict):
    try:
        clients = list(config.get("Clients").keys())
        client_map = {i: dirname for i, dirname in enumerate(clients, 1)}
        print(f"Found {len(clients)} clients in config file.")
        for i, client in client_map.items():
            print(f"{i} - {client}")

        for _ in range(3):
            choice = input("\nSelect client: ")
            logging.info(f"User's choice in 'select_client' function: {choice}")
            try:
                client = client_map[int(choice)]
            except ValueError:
                print("[WARNING] Please provide a number.\n")
                continue
            except KeyError:
                print("[WARNING] Please choose one of the numbers displayed. \n")
                continue
            else:
                break
        else:
            logging.warning("User entered three invald inputs. Script terminates.")
            input("Received three invalid inputs. Script termintates. Please press any key ...")
            sys.exit()

        client_info = config.get("Clients").get(client)
        logging.info("Successfully selected client.")
        return client, client_info
    
    except Exception as e:
        logging.info("Error in 'select_client' function. Terminated with error:")
        logging.error(f"{e}")
        sys.exit()

def handle_no_tasks(no_tasks: pd.DataFrame) -> None:
    try:
        print(f"Found {len(no_tasks)} entries without any task name.")
        print("Note, that these entries won't be included in the 'Stundenaufstellung'.")
        for _ in range(3):
            export = input("Would you like to export them in a separate excel file? (y/n) ")
            logging.info(f"User's choice in 'handle_no_tasks' function: {export}")
            if export == "y":
                wb = openpyxl.Workbook()
                ws = wb.active
                # add header
                ws.append(column_names)
                for _, row in no_tasks.iterrows():
                    ws.append(row.tolist())
                wb.save("no_tasks.xlsx")
                print("\nNew file created: 'no_tasks.xlsx'")
                print("However, I recommend adding tasks to your original Replicon Export.")
                print("Then, you can re-run the program to include all entries.")
                logging.info("Successfully handled no task entries.")
                return
            elif export == "n":
                logging.info("Successfully handled no task entries.")
                return
            else:
                print(f"'{export}' is not a valid input. Please select 'y' for 'yes' and 'n' for 'no'.")
        else:
            logging.warning("User entered three invald inputs. Script terminates.")
            input("Received three invalid inputs. Script termintates. Please press any key ...")
            sys.exit()

    except Exception as e:
        logging.info("Error in 'handle_no_tasks' function. Terminated with error:")
        logging.error(f"{e}")
        sys.exit()

def handle_unexpected_error(error):
    logging.info("Unexpected error in logic section. Terminated with error:")
    logging.error(f"{error}")
    print("[ERROR] Something unexpected happened.")
    print("[ERROR] Please, contact the developers and send them the log file.")
    input("[ERROR] Script terminates. Press any key to exit ...")
    sys.exit()
