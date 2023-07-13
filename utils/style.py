from copy import copy
from openpyxl.utils import range_boundaries
from utils.utils import add_logging


@add_logging
def copy_cell_styles(template_sheet, new_sheet):
    for i, row in enumerate(template_sheet.rows, start=1):
        for j, cell in enumerate(row, start=1):
            new_cell = new_sheet.cell(row=i, column=j, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)


# sheet is openpyxl worksheet
@add_logging
def adjust_cell_dimensions(sheet, target_dimensions: dict):
    for col, width in target_dimensions["widths"].items():
        sheet.column_dimensions[col].width = float(width)
    for row, height in target_dimensions["heights"].items():
        sheet.row_dimensions[row].height = float(height)

        
# sheet is openpyxl worksheet
# target_ranges are openpyxl merged cell ranges
@add_logging
def merge_cells(sheet, target_ranges: set):
    for merged_range in target_ranges:
        min_row, min_col, max_row, max_col = range_boundaries(merged_range.coord)
        sheet.merge_cells(start_row=min_col,
                            start_column=min_row,
                            end_row=max_col,
                            end_column=max_row)
