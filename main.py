import os
import shutil
import time
import calendar
import string
import locale
import logging

import openpyxl
import pandas as pd

from utils.interactive import get_input, handle_failed_input, select_client,\
                                handle_no_tasks, handle_unexpected_error
from utils.preprocessing import prepare_df, clean_name, merge_groups
from utils.report import ClientReport1, ClientReport2
from utils.style import adjust_cell_dimensions, copy_cell_styles, merge_cells
from utils.validation import validate_input_file
from utils.file_handling import create_folder, load_template, load_config


# to get dates and times in German format
locale.setlocale(locale.LC_ALL, "de_DE")

cwd = os.getcwd()

# load configurations
config = load_config(os.path.join(cwd, "Template", "config.yaml"))

# set up logging
logs_dir = config.get("logging").get("logs_dir", "logs")
log_level = config.get("logging").get("level", "info").upper()
os.makedirs(logs_dir, exist_ok=True)

log_file = os.path.join(logs_dir, "log_file.log")
logging.basicConfig(level=log_level, format="%(asctime)s [%(levelname)s] %(message)s",
                    filename=log_file, force=True)
logging.info("Execution started.")

CODE_TO_ACTIVITY = {
    "000": config["Categories"].get("000"),
    "001": config["Categories"].get("001"),
    "002": config["Categories"].get("002"),
    "003": config["Categories"].get("003"),
    "004": config["Categories"].get("004"),
    "005": config["Categories"].get("005"),
    "006": config["Categories"].get("006"),
    "007": config["Categories"].get("007"),
    "999": config["Categories"].get("999")
}
# codes where the user wants store additional information (comments, the text after a ':')
# after the corresponding code
ADDITIONAL_COMMENTS = config["Categories"].get("additional_comments_for_codes")

client, client_info = select_client(config)
file = get_input(cwd)

# check if the provided Replicon Export has english column names
try:
    validate_input_file(file)
except TypeError:
    logging.warning("User used non-german column names.")
    handle_failed_input()

# some preprocessing
data = prepare_df(file)

# get template excel sheet for styles
template, template_sheet, skip_style = load_template(
    os.path.join(cwd, "Template", f"template_{client}.xlsx"),
    client,
    config
)
if not skip_style:
    template_cell_dimensions = {
        "widths": {
            col: template_sheet.column_dimensions[col].width for col in string.ascii_uppercase
        },
        "heights": {
            row: template_sheet.row_dimensions[row].height for row in range(1,\
                                    client_info.get("max_range_rows"))\
                if template_sheet.row_dimensions[row].height is not None
        }
    }
    template_merged_cells = template_sheet.merged_cells.ranges

create_folder("output", os.path.join(cwd, "output"))

# get all available clients
client_groups = data.groupby(data["Client Name"])


if __name__ == "__main__":
    start = time.time()
    try:
        # client level (folder)
        for i, (client_name, group) in enumerate(client_groups, 1):
            create_folder(str(client_name), os.path.join(cwd, "output", str(client_name)))

            wbs_groups = group.groupby(group["Project Code"])
            # wbs code level (folder)
            for wbs, wbs_group in wbs_groups:
                project_name = data.loc[data["Project Code"] == wbs, "Project Name"].values[0]
                clean_project_name = clean_name(project_name)
                create_folder(f"{wbs} ({clean_project_name})")

                # -- CLIENT 1 --
                if client_info.get("id") == 1:
                    year_groups = wbs_group.groupby(wbs_group["Entry Date"].dt.year)
                    # year level (folder)
                    for year, year_group in year_groups:
                        create_folder(str(int(year)))
                        
                        month_groups = year_group.groupby(year_group["Entry Date"].dt.month)
                        # month level (excel file)
                        for month, month_group in month_groups:
                            shutil.copy(os.path.join(cwd,
                                                    "Template", f"template_{client}.xlsx"),
                                                    f"{calendar.month_name[month]}.xlsx")
                            month_excel = openpyxl.load_workbook(f"{calendar.month_name[month]}.xlsx")

                            employee_groups = month_group.groupby(month_group["Last Name"])
                            # employee level (excel sheet in the file)
                            # creates report for every employee
                            for employee, employee_group in employee_groups:
                                first_name = employee_group["First Name"].values[0]
                                employee_sheet = month_excel.create_sheet(f"{first_name} {employee}")
                                # call functions to copy style and values
                                merge_cells(employee_sheet, template_merged_cells)
                                copy_cell_styles(template_sheet, employee_sheet)
                                adjust_cell_dimensions(employee_sheet, template_cell_dimensions)
                                report = ClientReport1(employee_group,
                                                    month=month,
                                                    year=year,
                                                    employee_name=f"{first_name} {employee}",
                                                    project_name=clean_project_name,
                                                    references=client_info.get("references"),
                                                    header_references=client_info.get("header_references"))
                                report.fill_worksheet(employee_sheet, CODE_TO_ACTIVITY, ADDITIONAL_COMMENTS)
                                report.fill_header(employee_sheet)

                            month_excel.remove(month_excel["template"])
                            month_excel.save(f"{calendar.month_name[month]}.xlsx")
                            
                        os.chdir("..")
                    os.chdir("..")
                # -- END CLIENT 1 --

                # -- START CLIENT 2 --
                elif client_info.get("id") == 2:
                    shutil.copy(os.path.join(cwd,
                                            "Template", f"template_{client}.xlsx"),
                                            f"{client}_Stundenaufstellung.xlsx")
                    output_excel = openpyxl.load_workbook(f"{client}_Stundenaufstellung.xlsx")
                    
                    no_tasks = wbs_group[pd.isna(wbs_group["Task Name"])]
                    handle_no_tasks(no_tasks)

                    wbs_group.dropna(subset=["Task Name"], inplace=True)

                    task_name_groups = wbs_group.groupby(wbs_group["Task Name"])
                    # some tasks belong together, merge them and sort by date before
                    # creating the report
                    merged_groups = merge_groups(task_name_groups,
                                                client_info,
                                                long_task_name=True)

                    # create report for every task
                    for task_name, task_name_group in merged_groups.items():
                        task_sheet = output_excel[task_name]
                        report = ClientReport2(task_name_group,
                                            task_name=task_name,
                                            grades=config.get("Grades"),
                                            header_references=client_info.get("header_references"))
                        report.fill_worksheet(task_sheet, CODE_TO_ACTIVITY, ADDITIONAL_COMMENTS)
                        report.fill_header(output_excel["Uebersicht"])
                    
                    output_excel.save(f"{client}_Stundenaufstellung.xlsx")
                    # -- END CLIENT 2 --

                    # -- START CLIENT 3 --
                elif client_info.get("id") == 3:
                    shutil.copy(os.path.join(cwd,
                                            "Template", f"template_{client}.xlsx"),
                                            f"{client}_Stundenaufstellung.xlsx")
                    output_excel = openpyxl.load_workbook(f"{client}_Stundenaufstellung.xlsx")
                    
                    no_tasks = wbs_group[pd.isna(wbs_group["Task Name"])]
                    handle_no_tasks(no_tasks)

                    wbs_group.dropna(subset=["Task Name"], inplace=True)

                    task_name_groups = wbs_group.groupby(wbs_group["Task Name"])
                    # some tasks belong together, merge them and sort by date before
                    # creating the report             
                    merged_groups = merge_groups(task_name_groups,
                                                client_info,
                                                long_task_name=False)

                    # create report for every task
                    for task_name, task_name_group in merged_groups.items():
                        task_sheet = output_excel[task_name]
                        report = ClientReport2(task_name_group,
                                            task_name=task_name,
                                            grades=config.get("Grades"),
                                            header_references=client_info.get("header_references"))
                        report.fill_worksheet(task_sheet, CODE_TO_ACTIVITY, ADDITIONAL_COMMENTS)
                        report.fill_header(output_excel["Uebersicht"])
                    
                    output_excel.save(f"{client}_Stundenaufstellung.xlsx")
                    # -- END CLIENT 3 --

            os.chdir("..")
            print(f"Progress: {round((i/len(client_groups))*100, 2)} %", end="\r")
    
    except Exception as e:
        handle_unexpected_error(e)

    delta = round(time.time() - start, 3)

    input(f"Finished Execution in {delta} s. Press any key to exit ...")
